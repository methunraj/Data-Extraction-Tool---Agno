"""
Minimal FastAPI API layer for IntelliExtract.
Direct workflow usage with no custom abstractions - let Agno handle complexity.
"""

import os
import sys
import uuid
import asyncio
import json
import threading
import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
import aiofiles

# Load environment variables
from dotenv import load_dotenv

# Load from parent directory .env file
env_path = Path(__file__).parent.parent.parent / ".env"
load_dotenv(dotenv_path=env_path)

from fastapi import FastAPI, UploadFile, File, HTTPException, BackgroundTasks, Request
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

# Import our Agno-native workflows
from .workflows.data_transform import DataTransformWorkflow
from .workflows.prompt_engineer import PromptEngineerWorkflow, ExtractionSchema

# Import model factory functions
from .agents.models import get_extraction_model, get_transform_model

# Import settings and logging
from .core.config import settings
from .core.logging import setup_logging, get_logger, log_api_request, log_api_response, log_error, error_monitor
from .monitoring import metrics_collector, get_health_status
import uuid as uuid_module
import time

# Setup structured logging
setup_logging()
logger = get_logger("main")




# Pydantic models for API requests
class ConfigRequest(BaseModel):
    requirements: str
    sample_documents: Optional[List[str]] = None


class ExtractionRequest(BaseModel):
    request: str
    session_id: Optional[str] = None
    model_name: Optional[str] = None
    document_text: Optional[str] = None
    document_file: Optional[Dict[str, Any]] = None
    schema_definition: Optional[str] = None
    system_prompt: Optional[str] = None
    user_prompt_template: Optional[str] = None
    examples: Optional[List[Dict[str, str]]] = None


# Initialize FastAPI app with size limits
app = FastAPI(
    title="IntelliExtract API",
    description="AI-powered data extraction using Agno workflows",
    version="2.0.0",
)

# Add CORS middleware with environment-based configuration
cors_origins = settings.CORS_ORIGINS.split(",") if settings.CORS_ORIGINS else ["*"]
if settings.ENVIRONMENT == "production" and "*" in cors_origins:
    logger.warning("CORS configured to allow all origins in production. Consider restricting origins.")

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["*"],
)

# Add request logging and size limit middleware
@app.middleware("http")
async def logging_and_size_middleware(request: Request, call_next):
    """Middleware for request logging and size limits"""
    request_id = str(uuid_module.uuid4())
    start_time = time.time()
    
    # Log incoming request and record metrics
    log_api_request(
        request_id=request_id,
        method=request.method,
        path=request.url.path,
        query_params=str(request.query_params),
        user_agent=request.headers.get("user-agent", ""),
        client_ip=request.client.host if request.client else "unknown"
    )
    metrics_collector.record_request(request.method, request.url.path)
    
    # Check request size limits
    if request.method in ["POST", "PUT", "PATCH"]:
        content_length = request.headers.get("content-length")
        if content_length:
            content_length = int(content_length)
            max_size = settings.MAX_REQUEST_SIZE_MB * 1024 * 1024  # Convert MB to bytes
            if content_length > max_size:
                error_msg = f"Request too large. Maximum size is {settings.MAX_REQUEST_SIZE_MB}MB"
                log_error(
                    ValueError(error_msg),
                    context={
                        "request_id": request_id,
                        "content_length": content_length,
                        "max_size": max_size
                    }
                )
                return HTTPException(status_code=413, detail=error_msg)
    
    try:
        # Process request
        response = await call_next(request)
        
        # Log response and record metrics
        duration_ms = (time.time() - start_time) * 1000
        log_api_response(
            request_id=request_id,
            status_code=response.status_code,
            duration_ms=duration_ms
        )
        metrics_collector.record_response_time(duration_ms)
        
        return response
        
    except Exception as e:
        # Log error and record metrics
        duration_ms = (time.time() - start_time) * 1000
        log_error(e, context={
            "request_id": request_id,
            "method": request.method,
            "path": request.url.path,
            "duration_ms": duration_ms
        })
        metrics_collector.record_error(type(e).__name__)
        
        # Re-raise the exception
        raise



# Global settings
TEMP_DIR = os.environ.get("AGENT_TEMP_DIR", "/tmp/intelliextract")
Path(TEMP_DIR).mkdir(parents=True, exist_ok=True)

# Cancellation token storage
ACTIVE_OPERATIONS: Dict[str, threading.Event] = {}

class CancellationToken:
    """Simple cancellation token for long-running operations"""
    def __init__(self, operation_id: str):
        self.operation_id = operation_id
        self.cancelled = threading.Event()
        ACTIVE_OPERATIONS[operation_id] = self.cancelled
    
    def is_cancelled(self) -> bool:
        return self.cancelled.is_set()
    
    def cancel(self):
        self.cancelled.set()
    
    def cleanup(self):
        if self.operation_id in ACTIVE_OPERATIONS:
            del ACTIVE_OPERATIONS[self.operation_id]

@app.delete("/api/cancel-operation/{operation_id}")
async def cancel_operation(operation_id: str):
    """Cancel a running operation"""
    if operation_id in ACTIVE_OPERATIONS:
        ACTIVE_OPERATIONS[operation_id].set()
        return {"message": f"Cancellation requested for operation {operation_id}"}
    else:
        raise HTTPException(status_code=404, detail="Operation not found or already completed")


@app.post("/api/generate-config", response_model=Dict[str, Any])
async def generate_config(request: ConfigRequest):
    """
    Generate extraction configuration using PromptEngineerWorkflow.
    Direct workflow usage - no custom abstractions.
    """
    try:
        # Initialize workflow (lightweight - ~3μs in Agno)
        prompt_engineer = PromptEngineerWorkflow()

        # Single agent call returns structured data automatically
        if request.sample_documents:
            config: ExtractionSchema = prompt_engineer.run_with_examples(
                request.requirements, request.sample_documents
            )
        else:
            config: ExtractionSchema = prompt_engineer.run(request.requirements)

        # Return Pydantic model as dict - no custom serialization needed
        return config.model_dump()

    except Exception as e:
        # Minimal error handling - let FastAPI handle HTTP errors
        raise HTTPException(
            status_code=500, detail=f"Configuration generation failed: {str(e)}"
        )


@app.post("/api/generate-unified-config", response_model=Dict[str, Any])
async def generate_unified_config(request: Dict[str, Any], background_tasks: BackgroundTasks):
    """
    Generate unified extraction configuration (frontend-compatible endpoint).
    Maps frontend parameters to PromptEngineerWorkflow.
    """
    try:
        # Extract model_name from request
        model_name = request.get("model_name")
        print(f"[generate-unified-config] Received model_name: {model_name}")

        # Create cancellation token for this operation
        operation_id = str(uuid.uuid4())
        cancellation_token = CancellationToken(operation_id)
        
        try:
            # Initialize workflow with the selected model
            prompt_engineer = PromptEngineerWorkflow(model_id=model_name)

            # Map frontend parameters to backend format
            requirements = request.get("user_intent", "")
            
            # Validate and process sample_data
            sample_data_raw = request.get("sample_data", "")
            sample_documents = None
            
            if sample_data_raw:
                # Validate sample_data size (max 10MB)
                if len(sample_data_raw) > 10 * 1024 * 1024:
                    raise HTTPException(
                        status_code=400, 
                        detail="sample_data too large. Maximum size is 10MB"
                    )
                
                # Split and validate documents
                sample_docs = sample_data_raw.split("\n")
                
                # Filter out empty lines and validate document count
                valid_docs = [doc.strip() for doc in sample_docs if doc.strip()]
                
                if len(valid_docs) > 100:
                    raise HTTPException(
                        status_code=400, 
                        detail="Too many sample documents. Maximum is 100 documents"
                    )
                
                # Validate individual document size
                for i, doc in enumerate(valid_docs):
                    if len(doc) > 100000:  # 100KB per document
                        raise HTTPException(
                            status_code=400, 
                            detail=f"Sample document {i+1} too large. Maximum size per document is 100KB"
                        )
                
                sample_documents = valid_docs if valid_docs else None

            if not requirements:
                raise HTTPException(status_code=400, detail="user_intent is required")

            # Generate configuration
            if sample_documents and any(doc.strip() for doc in sample_documents):
                config: ExtractionSchema = prompt_engineer.run_with_examples(
                    requirements, [doc.strip() for doc in sample_documents if doc.strip()]
                )
            else:
                config: ExtractionSchema = prompt_engineer.run(requirements)
            
            # Check for cancellation after generation
            if cancellation_token.is_cancelled():
                raise HTTPException(status_code=499, detail="Operation was cancelled")
                
        except Exception as e:
            log_error(e, context={
                "endpoint": "generate-unified-config",
                "user_intent": requirements[:100] if requirements else None,
                "has_sample_data": bool(sample_documents)
            })
            raise HTTPException(status_code=500, detail=f"Configuration generation failed: {str(e)}")
        finally:
            # Always cleanup the cancellation token
            cancellation_token.cleanup()

        # Convert to frontend format
        result = config.model_dump()

        # Add additional fields expected by frontend
        # Convert Example objects to dicts if needed
        examples = result.get("examples", [])
        if examples and hasattr(examples[0], "model_dump"):
            examples = [ex.model_dump() for ex in examples]

        return {
            "schema": result.get("json_schema", "{}"),
            "system_prompt": result.get("system_prompt", ""),
            "user_prompt_template": result.get("user_prompt_template", ""),
            "examples": examples,
            "reasoning": "Configuration generated successfully using Agno PromptEngineerWorkflow",
            "cost": 0.001,  # Placeholder cost
            "tokens_used": 1000,  # Placeholder token count
            "operation_id": operation_id,
        }

    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Unified configuration generation failed: {str(e)}"
        )


@app.post("/api/refine-config", response_model=Dict[str, Any])
async def refine_config(current_config: Dict[str, Any], feedback: str):
    """
    Refine existing configuration based on feedback.
    """
    try:
        prompt_engineer = PromptEngineerWorkflow()

        # Convert dict back to Pydantic model
        config_obj = ExtractionSchema(**current_config)

        # Refine configuration
        refined_config = prompt_engineer.refine_configuration(config_obj, feedback)
        return refined_config.model_dump()

    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Configuration refinement failed: {str(e)}"
        )


@app.post("/api/extract-data")
async def extract_data_json(request: Dict[str, Any], background_tasks: BackgroundTasks):
    """
    Extract data endpoint that accepts JSON input (frontend-compatible).
    Uses Agno to process documents directly.
    """
    try:
        import os
        from agno.agent import Agent
        from agno.models.google import Gemini

        # Extract parameters from request
        document_text = request.get("document_text", "")
        document_file = request.get("document_file")
        schema_definition = request.get("schema_definition", "{}")
        system_prompt = request.get(
            "system_prompt", "Extract data according to the schema"
        )
        user_prompt_template = request.get("user_prompt_template", "{document_text}")
        examples = request.get("examples", [])
        model_name = request.get("model_name", "gemini-2.0-flash")
        print(f"[extract-data] Received model_name: {model_name}")
        print(f"[extract-data] Has document_text: {bool(document_text)}")
        print(f"[extract-data] Has document_file: {bool(document_file)}")
        if document_file:
            print(f"[extract-data] File mime_type: {document_file.get('mime_type')}")
            print(
                f"[extract-data] File data length: {len(document_file.get('data', ''))}"
            )

        # Handle document content
        content = document_text
        files = []
        tmp_file_path = None

        if not content and document_file:
            # Handle PDF/image files with Agno's multimodal capabilities
            import base64
            import tempfile
            from agno.media import File

            mime_type = document_file.get("mime_type", "application/pdf")
            file_data = document_file.get("data", "")

            try:
                # Validate file size before processing to prevent memory exhaustion
                estimated_size = len(file_data) * 3 // 4  # Rough base64 to binary size estimate
                max_file_size = 100 * 1024 * 1024  # 100MB limit
                
                if estimated_size > max_file_size:
                    raise HTTPException(
                        status_code=413, 
                        detail=f"File too large. Maximum size is {max_file_size // (1024*1024)}MB"
                    )
                
                # Create a temporary file for the PDF
                with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_file:
                    # Decode entire base64 data and write to file
                    file_bytes = base64.b64decode(file_data)
                    tmp_file.write(file_bytes)
                    tmp_file_path = tmp_file.name

                # Create Agno File object
                files = [File(filepath=tmp_file_path)]
                content = "Please extract data from the attached PDF document."
                
            except Exception as e:
                # Clean up temp file if creation failed
                if tmp_file_path and os.path.exists(tmp_file_path):
                    try:
                        os.unlink(tmp_file_path)
                    except:
                        pass
                raise HTTPException(
                    status_code=500, 
                    detail=f"Failed to process document file: {str(e)}"
                )

        # Create extraction agent with the selected model
        agent = Agent(
            name="DataExtractor",
            model=Gemini(id=model_name, api_key=os.environ.get("GOOGLE_API_KEY")),
            instructions=[system_prompt],
            markdown=False,
            show_tool_calls=False,
        )

        # Build the extraction prompt
        if files:
            # For PDF/image files, adjust the prompt
            prompt = system_prompt + "\n\n"
            prompt += user_prompt_template.replace(
                "{document_text}", "the attached PDF document"
            )
            prompt += (
                f"\n\nExtract data according to this JSON schema:\n{schema_definition}"
            )
        else:
            # For text content
            prompt = user_prompt_template.replace("{document_text}", content)
            prompt += (
                f"\n\nExtract data according to this JSON schema:\n{schema_definition}"
            )

        if examples:
            prompt += "\n\nExamples:"
            for ex in examples:
                prompt += (
                    f"\nInput: {ex.get('input', '')}\nOutput: {ex.get('output', '')}\n"
                )

        # Create cancellation token for this operation
        operation_id = str(uuid.uuid4())
        cancellation_token = CancellationToken(operation_id)
        
        try:
            # Run extraction with cancellation support
            if files:
                response = agent.run(prompt, files=files)
            else:
                response = agent.run(prompt)
            
            # Check for cancellation after agent run
            if cancellation_token.is_cancelled():
                raise HTTPException(status_code=499, detail="Operation was cancelled")

            # Extract the content
            extracted_json = ""
            if hasattr(response, "content"):
                extracted_json = response.content
                # Try to parse as JSON if it's a string
                if isinstance(extracted_json, str):
                    try:
                        import json

                        extracted_json = json.loads(extracted_json)
                    except:
                        # If it fails, keep as string
                        pass
                        
        finally:
            # Always cleanup the cancellation token
            cancellation_token.cleanup()
            
            # Always clean up temporary file if created
            if tmp_file_path and os.path.exists(tmp_file_path):
                try:
                    os.unlink(tmp_file_path)
                    print(f"[extract-data] Cleaned up temporary file: {tmp_file_path}")
                except Exception as cleanup_error:
                    print(f"[extract-data] Failed to cleanup temp file {tmp_file_path}: {cleanup_error}")

        # Return response in expected format
        return {
            "extracted_json": (
                json.dumps(extracted_json)
                if not isinstance(extracted_json, str)
                else extracted_json
            ),
            "token_usage": {
                "prompt_tokens": 1000,  # Placeholder
                "completion_tokens": 500,  # Placeholder
                "total_tokens": 1500,  # Placeholder
                "cached_tokens": 0,
                "thinking_tokens": 0,
            },
            "cost": 0.001,  # Placeholder
            "model_used": model_name,
            "cache_hit": False,
            "retry_count": 0,
            "thinking_text": None,
        }

    except Exception as e:
        import traceback

        error_detail = f"Extraction failed: {str(e)}\n{traceback.format_exc()}"
        print(error_detail)  # Log to console
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/extract-data-old")
async def extract_data(
    request: ExtractionRequest,
    files: List[UploadFile] = File(...),
    background_tasks: BackgroundTasks = None,
):
    """
    Extract data and generate Excel report with streaming responses.
    Uses Agno's native streaming - Iterator[RunResponse] -> SSE.
    """

    # Create unique session directory
    session_id = request.session_id or str(uuid.uuid4())
    session_dir = os.path.join(TEMP_DIR, f"session_{session_id}")
    Path(session_dir).mkdir(parents=True, exist_ok=True)

    # Save uploaded files to session directory
    saved_files = []
    try:
        for file in files:
            # Validate file size
            content = await file.read()
            file_size_mb = len(content) / (1024 * 1024)
            if file_size_mb > settings.MAX_FILE_SIZE_MB:
                raise HTTPException(
                    status_code=413,
                    detail=f"File '{file.filename}' is too large ({file_size_mb:.1f}MB). Maximum size is {settings.MAX_FILE_SIZE_MB}MB"
                )
            
            file_path = os.path.join(
                session_dir, file.filename or f"file_{uuid.uuid4().hex[:8]}"
            )
            async with aiofiles.open(file_path, "wb") as f:
                await f.write(content)
            saved_files.append(
                {
                    "name": file.filename,
                    "path": file_path,
                    "type": file.content_type,
                    "size": len(content),
                }
            )

        # Initialize workflow with session directory and model
        data_transform = DataTransformWorkflow(
            working_dir=session_dir, model_id=request.model_name
        )

        # Stream workflow responses using Server-Sent Events
        def stream_responses():
            """
            Stream Agno workflow responses as SSE.
            Direct Iterator[RunResponse] -> SSE conversion.
            """
            try:
                # Run workflow with streaming - pure Python iterator
                for response in data_transform.run(request.request, saved_files):
                    # Convert RunResponse to SSE format
                    if hasattr(response, "content") and response.content:
                        # SSE format: data: {content}\n\n
                        yield f"data: {response.content}\n\n"

                # Send completion event
                yield f'data: {{"status": "completed", "session_id": "{session_id}"}}\n\n'

            except Exception as e:
                # Stream error as SSE
                yield f'data: {{"error": "Workflow failed: {str(e)}"}}\n\n'

        # Schedule cleanup as background task
        if background_tasks:
            background_tasks.add_task(
                cleanup_session, session_dir, delay_seconds=300
            )  # 5 minutes

        # Return streaming response
        return StreamingResponse(
            stream_responses(),
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "Connection": "keep-alive",
                "X-Session-ID": session_id,
            },
        )

    except Exception as e:
        # Cleanup on immediate failure
        cleanup_session_sync(session_dir)
        raise HTTPException(status_code=500, detail=f"Data extraction failed: {str(e)}")


@app.get("/api/download-report/{session_id}")
async def download_report(session_id: str, filename: Optional[str] = None):
    """
    Download generated Excel report from session directory.
    """
    import re
    
    # Validate session_id to prevent directory traversal
    if not re.match(r'^[a-zA-Z0-9_-]+$', session_id):
        raise HTTPException(
            status_code=400, detail="Invalid session ID format"
        )
    
    session_dir = os.path.join(TEMP_DIR, f"session_{session_id}")
    
    # Ensure session directory exists and is within TEMP_DIR
    if not os.path.exists(session_dir) or not os.path.commonpath([TEMP_DIR, session_dir]) == TEMP_DIR:
        raise HTTPException(
            status_code=404, detail="Session not found"
        )

    # Look for Excel files in session directory
    excel_files = list(Path(session_dir).glob("*.xlsx"))

    if not excel_files:
        raise HTTPException(
            status_code=404, detail="No Excel report found for this session"
        )

    # Use specified filename or first found Excel file
    if filename:
        # Sanitize filename to prevent path traversal
        safe_filename = os.path.basename(filename)
        if not safe_filename or safe_filename != filename or '..' in filename:
            raise HTTPException(
                status_code=400, detail="Invalid filename"
            )
        
        # Ensure filename has valid extension
        if not safe_filename.endswith('.xlsx'):
            raise HTTPException(
                status_code=400, detail="Only .xlsx files are allowed"
            )
        
        report_path = os.path.join(session_dir, safe_filename)
        
        # Double-check the resolved path is still within session directory
        if not os.path.commonpath([session_dir, report_path]) == session_dir:
            raise HTTPException(
                status_code=400, detail="Invalid file path"
            )
        
        if not os.path.exists(report_path):
            raise HTTPException(status_code=404, detail=f"File {safe_filename} not found")
    else:
        report_path = str(excel_files[0])
        safe_filename = os.path.basename(report_path)

    # Return file response
    return FileResponse(
        path=report_path,
        filename=safe_filename or "extracted_data.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/api/agno-process")
async def agno_process(request: Request, background_tasks: BackgroundTasks):
    """
    Process extracted data with Agno to generate Excel file.
    """
    try:
        # Parse request body
        body = await request.json()
        extracted_data = body.get("extractedData")
        file_name = body.get("fileName", "output.xlsx")
        model_name = body.get("model", "gemini-2.0-flash")

        if not extracted_data:
            raise HTTPException(status_code=400, detail="Missing extractedData")

        # Parse extracted data if it's a string
        if isinstance(extracted_data, str):
            # Check if it's markdown-wrapped JSON
            if extracted_data.strip().startswith('```json') and extracted_data.strip().endswith('```'):
                # Remove markdown wrapper
                print("[agno-process] Detected markdown-wrapped JSON, removing wrapper")
                extracted_data = extracted_data.strip()[7:-3].strip()  # Remove ```json and ```
            
            try:
                extracted_data = json.loads(extracted_data)
            except json.JSONDecodeError as e:
                print(f"[agno-process] Failed to parse JSON: {e}")
                # If it fails to parse, keep as string
                pass

        # Create session directory
        session_id = str(uuid.uuid4())
        session_dir = os.path.join(TEMP_DIR, f"session_{session_id}")
        os.makedirs(session_dir, exist_ok=True)

        # Save extracted data as JSON file for processing
        json_file_path = os.path.join(session_dir, "extracted_data.json")
        try:
            # Debug: Check data type before saving
            print(f"[agno-process] Data type before saving: {type(extracted_data)}")
            if isinstance(extracted_data, dict):
                print(f"[agno-process] Data is a dict with keys: {list(extracted_data.keys())[:5]}")
            
            with open(json_file_path, "w", encoding='utf-8') as f:
                json.dump(extracted_data, f, indent=2, ensure_ascii=False, default=str)
        except (TypeError, ValueError) as e:
            # Handle non-serializable objects by converting to string representation
            print(f"[agno-process] JSON serialization failed, using fallback: {e}")
            try:
                # Attempt to serialize with custom handler
                import pickle
                import base64
                
                def json_serializer(obj):
                    """Custom JSON serializer for complex objects"""
                    if hasattr(obj, '__dict__'):
                        return obj.__dict__
                    elif hasattr(obj, 'model_dump'):  # Pydantic models
                        return obj.model_dump()
                    elif hasattr(obj, 'dict'):  # Pydantic v1 models
                        return obj.dict()
                    else:
                        return str(obj)
                
                with open(json_file_path, "w", encoding='utf-8') as f:
                    json.dump(extracted_data, f, indent=2, ensure_ascii=False, default=json_serializer)
            except Exception as fallback_error:
                print(f"[agno-process] Fallback serialization failed: {fallback_error}")
                # Last resort: save as string representation
                with open(json_file_path, "w", encoding='utf-8') as f:
                    json.dump({"data": str(extracted_data), "error": "Original data was not JSON serializable"}, f, indent=2)
        except Exception as e:
            print(f"[agno-process] File write error: {e}")
            raise HTTPException(
                status_code=500, 
                detail=f"Failed to save extracted data: {str(e)}"
            )

        # Create cancellation token for this operation
        operation_id = str(uuid.uuid4())
        cancellation_token = CancellationToken(operation_id)
        
        # Initialize Excel generation workflow
        from .agents.excel_generator import ExcelGeneratorAgent
        excel_agent = ExcelGeneratorAgent(
            model=get_transform_model(model_id=model_name), working_dir=session_dir
        )
        print(f"[agno-process] Using ExcelGeneratorAgent with operation_id: {operation_id}")

        # Generate Excel file with clear, step-by-step prompt
        excel_prompt = f"""Convert the JSON data at {json_file_path} to an Excel file at {os.path.join(session_dir, file_name)}.

Follow these steps:

1. Install required packages:
   - pip_install_package pandas
   - pip_install_package openpyxl

2. Read and analyze the JSON:
   - Use read_file to read {json_file_path}
   - Understand the data structure

3. Write a Python script that:
   - Reads the JSON file
   - Creates an Excel workbook with pandas
   - Creates separate sheets for different data categories
   - Handles nested structures properly
   - Saves to {os.path.join(session_dir, file_name)}

4. Execute your script using save_to_file_and_run

5. Verify the Excel file was created successfully

Remember: The JSON contains financial data with nested structures. Create appropriate sheets for company info, financial metrics, and metadata."""

        # Run Excel generation
        print(f"[agno-process] Running Excel generation for session {session_id}")
        print(f"[agno-process] Working directory: {session_dir}")
        print(f"[agno-process] Expected output file: {file_name}")
        
        # Run Excel generation with proper continuation handling
        print(f"[agno-process] Starting Excel generation for session {session_id}")
        
        try:
            import time
            
            # Add timeout mechanism
            start_time = time.time()
            max_execution_time = 600  # 10 minutes timeout
            
            # Initial run
            result = excel_agent.run(excel_prompt)
            print(f"[agno-process] Initial run completed. Type: {type(result)}")
            
            # Handle continuation using the proper pattern from documentation
            max_continuations = 10
            continuation_count = 0
            
            while continuation_count < max_continuations:
                # Check for timeout first
                elapsed_time = time.time() - start_time
                if elapsed_time > max_execution_time:
                    print(f"[agno-process] Operation timed out after {elapsed_time:.1f} seconds")
                    raise HTTPException(
                        status_code=408, 
                        detail=f"Excel generation timed out after {max_execution_time} seconds"
                    )
                
                # Check for cancellation
                if cancellation_token.is_cancelled():
                    print("[agno-process] Operation cancelled by user")
                    raise HTTPException(status_code=499, detail="Operation was cancelled")
                
                # Check if agent is paused (preferred method)
                if hasattr(excel_agent, 'is_paused') and excel_agent.is_paused:
                    print("[agno-process] Agent is paused, continuing...")
                    result = excel_agent.continue_run()
                    continuation_count += 1
                    continue
                
                # Check status if available
                if hasattr(result, 'status'):
                    print(f"[agno-process] Run status: {result.status}")
                    
                    if result.status == "PAUSED":
                        print("[agno-process] Status is PAUSED, continuing...")
                        result = excel_agent.continue_run(run_response=result)
                        continuation_count += 1
                        continue
                    elif result.status in ["COMPLETED", "CANCELLED"]:
                        print(f"[agno-process] Run {result.status}")
                        break
                
                # Check if Excel file was created
                excel_path = os.path.join(session_dir, file_name)
                if os.path.exists(excel_path):
                    print(f"[agno-process] Success! Excel file created at: {excel_path}")
                    break
                
                # If no pause/status but no file, try one more continuation
                if continuation_count == 0:
                    print("[agno-process] No file created yet, attempting continuation...")
                    try:
                        # Check timeout and cancellation before continuing
                        elapsed_time = time.time() - start_time
                        if elapsed_time > max_execution_time:
                            print(f"[agno-process] Timeout reached during continuation check")
                            break
                        if cancellation_token.is_cancelled():
                            print("[agno-process] Operation cancelled during continuation")
                            break
                        result = excel_agent.continue_run()
                        continuation_count += 1
                    except Exception as e:
                        print(f"[agno-process] Continue failed: {e}")
                        break
                else:
                    # No more continuations
                    break
            
            # Final check for Excel file
            excel_path = os.path.join(session_dir, file_name)
            if not os.path.exists(excel_path):
                # Check for any Excel files
                excel_files = list(Path(session_dir).glob("*.xlsx"))
                if excel_files:
                    excel_path = str(excel_files[0])
                    file_name = excel_files[0].name
                    print(f"[agno-process] Found Excel file: {excel_path}")
                else:
                    all_files = list(Path(session_dir).iterdir())
                    print(f"[agno-process] No Excel file created. Files: {[f.name for f in all_files]}")
                    raise HTTPException(
                        status_code=500,
                        detail="Excel generation failed - no output file created"
                    )
                    
        except Exception as e:
            print(f"[agno-process] Excel generation error: {str(e)}")
            raise HTTPException(
                status_code=500,
                detail=f"Excel generation failed: {str(e)}"
            )

        # After all attempts, check if Excel file was created
        excel_path = os.path.join(session_dir, file_name)
        print(f"[agno-process] Final check for Excel file at: {excel_path}")
        
        if not os.path.exists(excel_path):
            # Try to find any Excel file in the directory
            excel_files = list(Path(session_dir).glob("*.xlsx"))
            print(f"[agno-process] Excel files found in directory: {[str(f) for f in excel_files]}")
            
            if excel_files:
                excel_path = str(excel_files[0])
                file_name = excel_files[0].name
                print(f"[agno-process] Using found Excel file: {excel_path}")
            else:
                # List all files in the directory for debugging
                all_files = list(Path(session_dir).iterdir())
                print(f"[agno-process] All files in directory: {[str(f) for f in all_files]}")
                
                # One final attempt with explicit instructions
                print("[agno-process] Making final attempt with explicit script...")
                final_prompt = f"""No Excel file was created. Let me provide you with a complete solution.

Use save_to_file_and_run with this exact script:

```python
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Read JSON
with open('{json_file_path}', 'r') as f:
    data = json.load(f)

# Create workbook
wb = Workbook()
wb.remove(wb.active)

# Sheet 1: Company Info
ws1 = wb.create_sheet("Company Information")
if 'company_identification' in data:
    for key, value in data['company_identification'].items():
        ws1.append([key.replace('_', ' ').title(), str(value)])

# Sheet 2: Financial Metrics
if 'financial_metrics' in data:
    for metric_type, metric_list in data['financial_metrics'].items():
        if isinstance(metric_list, list) and metric_list:
            ws = wb.create_sheet(metric_type.replace('_', ' ').title()[:31])
            if metric_list:
                headers = list(metric_list[0].keys())
                ws.append(headers)
                for item in metric_list:
                    ws.append([item.get(h, '') for h in headers])

# Save
wb.save('{excel_path}')
print(f"Excel file created at: {excel_path}")
```

Execute this script now."""
                
                try:
                    final_result = excel_agent.run(final_prompt)
                    print(f"[agno-process] Final attempt result: {final_result}")
                    
                    # Check one more time
                    if os.path.exists(excel_path):
                        print(f"[agno-process] Success! Excel file created at: {excel_path}")
                    else:
                        raise HTTPException(
                            status_code=500, detail="Excel file generation failed - no .xlsx file created"
                        )
                except Exception as e:
                    print(f"[agno-process] Final attempt failed: {str(e)}")
                    raise HTTPException(
                        status_code=500, detail="Excel file generation failed - no .xlsx file created"
                    )

        # Schedule cleanup
        background_tasks.add_task(
            cleanup_session, session_dir, delay_seconds=3600
        )  # 1 hour
        
        # Cleanup cancellation token
        cancellation_token.cleanup()

        # Return download URL
        return {
            "success": True,
            "download_url": f"/api/download-report/{session_id}?filename={file_name}",
            "file_name": file_name,
            "session_id": session_id,
            "operation_id": operation_id,
        }

    except Exception as e:
        # Cleanup cancellation token on error
        if 'cancellation_token' in locals():
            cancellation_token.cleanup()
        raise HTTPException(status_code=500, detail=f"Agno processing failed: {str(e)}")


@app.post("/api/test-excel-generation")
async def test_excel_generation():
    """
    Test endpoint for Excel generation to debug issues.
    """
    try:
        # Create test data
        test_data = {
            "products": [
                {"id": 1, "name": "Product A", "price": 100.50, "quantity": 10},
                {"id": 2, "name": "Product B", "price": 200.75, "quantity": 5},
                {"id": 3, "name": "Product C", "price": 50.25, "quantity": 20}
            ]
        }
        
        # Create session directory
        session_id = str(uuid.uuid4())
        session_dir = os.path.join(TEMP_DIR, f"test_{session_id}")
        os.makedirs(session_dir, exist_ok=True)
        
        # Save test data
        json_file_path = os.path.join(session_dir, "test_data.json")
        with open(json_file_path, "w") as f:
            json.dump(test_data, f, indent=2)
        
        # Initialize Excel generator
        from .agents.excel_generator import ExcelGeneratorAgent
        excel_agent = ExcelGeneratorAgent(
            model=get_transform_model(model_id="gemini-2.0-flash"),
            working_dir=session_dir
        )
        
        # Simple prompt
        excel_prompt = f"""
        You need to create an Excel file from JSON data using Python.
        
        Input file: {json_file_path}
        Output file: {os.path.join(session_dir, "test_output.xlsx")}
        
        Steps:
        1. First install pandas and openpyxl using pip_install_package
        2. Write a Python script that reads the JSON and creates an Excel file
        3. Use save_to_file_and_run to execute your script
        
        The JSON contains product data with id, name, price, and quantity fields.
        """
        
        # Run generation
        result = excel_agent.run(excel_prompt)
        
        # Check results
        excel_files = list(Path(session_dir).glob("*.xlsx"))
        
        return {
            "success": True,
            "session_id": session_id,
            "session_dir": session_dir,
            "excel_files": [str(f) for f in excel_files],
            "all_files": [str(f) for f in Path(session_dir).iterdir()],
            "agent_result": str(result.content) if hasattr(result, 'content') else str(result)
        }
        
    except Exception as e:
        import traceback
        return {
            "success": False,
            "error": str(e),
            "traceback": traceback.format_exc()
        }


@app.get("/api/sessions/{session_id}/files")
async def list_session_files(session_id: str):
    """
    List files in a session directory.
    """
    session_dir = os.path.join(TEMP_DIR, f"session_{session_id}")

    if not os.path.exists(session_dir):
        raise HTTPException(status_code=404, detail="Session not found")

    files = []
    for file_path in Path(session_dir).iterdir():
        if file_path.is_file():
            files.append(
                {
                    "name": file_path.name,
                    "size": file_path.stat().st_size,
                    "type": (
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        if file_path.suffix == ".xlsx"
                        else "unknown"
                    ),
                }
            )

    return {"session_id": session_id, "files": files}


@app.delete("/api/sessions/{session_id}")
async def cleanup_session_endpoint(session_id: str, background_tasks: BackgroundTasks):
    """
    Manually cleanup a session directory.
    """
    session_dir = os.path.join(TEMP_DIR, f"session_{session_id}")

    if not os.path.exists(session_dir):
        raise HTTPException(status_code=404, detail="Session not found")

    # Schedule cleanup as background task
    background_tasks.add_task(cleanup_session_sync, session_dir)

    return {"message": f"Session {session_id} cleanup scheduled"}


@app.get("/api/models")
async def get_models():
    """
    Get available models from models.json configuration.
    """
    try:
        # Import here to avoid circular dependency
        from .services.model_service import get_model_service

        model_service = get_model_service()
        models = model_service.get_all_models()

        # Convert to frontend-compatible format
        formatted_models = []
        for model in models:
            formatted_models.append(
                {
                    "id": model["id"],
                    "displayName": model["displayName"],
                    "description": model["description"],
                    "provider": model["provider"],
                    "supportedIn": model["supportedIn"],
                    "capabilities": model.get("capabilities", {}),
                    "limits": model.get("limits", {}),
                    "pricing": model.get("pricing", {}),
                    "status": model.get("status", "stable"),
                }
            )

        return {"models": formatted_models}
    except Exception as e:
        # Fallback to minimal hardcoded models if service fails
        print(f"Error loading models from service: {e}")
        return {
            "models": [
                {
                    "id": "gemini-2.0-flash-001",
                    "displayName": "Gemini 2.0 Flash",
                    "description": "Fast model for extraction",
                    "provider": "googleAI",
                    "supportedIn": ["extraction", "generation", "agno"],
                    "capabilities": {"vision": True},
                    "limits": {"maxInputTokens": 1000000},
                    "pricing": {"input": 0.0001, "output": 0.0004},
                    "status": "stable",
                }
            ]
        }


@app.get("/health")
async def health():
    """
    Health check endpoint with comprehensive monitoring.
    """
    health_status = get_health_status()
    
    # Add basic service info
    health_status.update({
        "framework": "agno",
        "api_version": "2.0.0",
        "temp_dir": TEMP_DIR,
        "temp_dir_exists": os.path.exists(TEMP_DIR),
        "environment": settings.ENVIRONMENT,
        "python_version": sys.version.split()[0],
        "platform": sys.platform
    })
    
    return health_status


@app.get("/api/monitoring/metrics")
async def get_monitoring_metrics():
    """
    Get detailed monitoring metrics.
    """
    return metrics_collector.get_metrics()


@app.get("/api/monitoring/errors")
async def get_error_monitoring():
    """
    Get error monitoring statistics.
    """
    return {
        "error_stats": error_monitor.get_error_stats(),
        "total_errors": sum(error_monitor.get_error_stats().values()),
        "timestamp": datetime.datetime.utcnow().isoformat() + "Z"
    }


@app.get("/")
async def root():
    """
    Root endpoint.
    """
    return {
        "message": "IntelliExtract API - AI-powered data extraction using Agno workflows",
        "docs": "/docs",
        "health": "/health",
    }


# Utility functions for cleanup
async def cleanup_session(session_dir: str, delay_seconds: int = 0):
    """
    Async cleanup function for background tasks.
    """
    if delay_seconds > 0:
        await asyncio.sleep(delay_seconds)

    cleanup_session_sync(session_dir)


def cleanup_session_sync(session_dir: str):
    """
    Synchronous cleanup function.
    """
    try:
        import shutil

        if os.path.exists(session_dir):
            shutil.rmtree(session_dir)
            print(f"Cleaned up session directory: {session_dir}")
    except Exception as e:
        print(f"Failed to cleanup session directory {session_dir}: {e}")

@app.post("/api/generate-schema")
async def generate_schema(request: Dict[str, Any]):
    """
    Generate a JSON schema based on user intent.
    """
    try:
        intent = request.get("intent", "")
        if not intent:
            raise HTTPException(status_code=400, detail="Intent is required")
        
        # Initialize prompt engineer workflow
        prompt_engineer = PromptEngineerWorkflow(
            model=get_extraction_model(model_id=settings.DEFAULT_AI_MODEL)
        )
        
        # Generate schema using the prompt engineer
        schema_prompt = f"""Generate a JSON schema for data extraction based on this intent: {intent}

Please create a comprehensive JSON schema that includes:
1. Appropriate field names and types
2. Required vs optional fields
3. Field descriptions
4. Validation constraints where applicable

Return only the JSON schema without any markdown formatting."""
        
        result = prompt_engineer.run(schema_prompt)
        
        # Extract schema from result
        if hasattr(result, 'content'):
            schema_content = result.content
        else:
            schema_content = str(result)
        
        # Try to parse as JSON to validate
        try:
            import json
            parsed_schema = json.loads(schema_content)
            return {"schema": schema_content, "success": True}
        except json.JSONDecodeError:
            # If not valid JSON, return as-is but mark as potentially invalid
            return {"schema": schema_content, "success": True, "warning": "Generated schema may need manual validation"}
            
    except Exception as e:
        logger.error(f"Schema generation failed: {e}")
        raise HTTPException(status_code=500, detail=f"Schema generation failed: {str(e)}")


# App startup event
@app.on_event("startup")
async def startup_event():
    """
    Application startup - ensure temp directory exists.
    """
    Path(TEMP_DIR).mkdir(parents=True, exist_ok=True)
    print(f"IntelliExtract API started - Temp directory: {TEMP_DIR}")


# App shutdown event
@app.on_event("shutdown")
async def shutdown_event():
    """
    Application shutdown - optional cleanup.
    """
    print("IntelliExtract API shutting down")


if __name__ == "__main__":
    import uvicorn

    port = int(os.getenv("PORT", 8000))
    uvicorn.run(
        "app.main:app", host="0.0.0.0", port=port, reload=True, log_level="info"
    )
