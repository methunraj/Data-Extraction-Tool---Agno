# app/agents/codegen_agent.py
import os
import logging
from pathlib import Path
from typing import Optional, Dict

from agno.agent import Agent
from agno.tools.python import PythonTools
from agno.tools.shell import ShellTools
from agno.storage.sqlite import SqliteStorage

from ..base import BaseAgent
from ...core.config import settings

logger = logging.getLogger(__name__)


class CodeGenAgent(BaseAgent):
    """Python execution agent for Excel generation with comprehensive tools."""
    
    def __init__(self, temp_dir: str, exchange_rates: Optional[Dict[str, float]] = None, model_id=None):
        super().__init__("codegen", temp_dir, model_id=model_id)
        self.exchange_rates = exchange_rates
        if not temp_dir:
            raise ValueError("temp_dir is required for codegen agent")
        
        # Validate environment on initialization
        if not self.validate_environment():
            raise RuntimeError(f"Environment validation failed for temp_dir: {temp_dir}")
    
    def validate_environment(self) -> bool:
        """Validate the execution environment before running"""
        try:
            # Ensure directory exists
            os.makedirs(self.temp_dir, exist_ok=True)
            
            # Test file creation
            test_file = os.path.join(self.temp_dir, '.test_write')
            with open(test_file, 'w') as f:
                f.write('test')
            
            # Test file reading
            with open(test_file, 'r') as f:
                content = f.read()
                if content != 'test':
                    raise ValueError("File read/write test failed")
            
            # Clean up test file
            os.remove(test_file)
            
            # Check permissions
            if not os.access(self.temp_dir, os.W_OK):
                raise PermissionError(f"No write access to {self.temp_dir}")
            
            if not os.access(self.temp_dir, os.R_OK):
                raise PermissionError(f"No read access to {self.temp_dir}")
            
            logger.info(f"Environment validation passed for: {self.temp_dir}")
            return True
            
        except Exception as e:
            logger.error(f"Environment validation failed: {e}")
            return False
    
    def get_file_operations_code(self) -> str:
        """Generate reliable file operation code"""
        abs_temp_dir = os.path.abspath(self.temp_dir)
        return f'''
import os
import sys
import shutil

class FileOperations:
    OUTPUT_DIR = r'{abs_temp_dir}'
    
    @classmethod
    def save_excel(cls, wb, filename='financial_report.xlsx'):
        """Reliably save Excel file to correct location"""
        os.makedirs(cls.OUTPUT_DIR, exist_ok=True)
        filepath = os.path.join(cls.OUTPUT_DIR, filename)
        
        # Try primary save
        try:
            wb.save(filepath)
            print(f'✅ Saved to: {{filepath}}')
            if cls.verify_file(filepath):
                return filepath
        except Exception as e:
            print(f'⚠️ Primary save failed: {{e}}')
        
        # Fallback: try alternative filename
        try:
            alt_filename = f'backup_{{filename}}'
            alt_filepath = os.path.join(cls.OUTPUT_DIR, alt_filename)
            wb.save(alt_filepath)
            print(f'✅ Saved via fallback to: {{alt_filepath}}')
            if cls.verify_file(alt_filepath):
                return alt_filepath
        except Exception as e:
            print(f'⚠️ Fallback save failed: {{e}}')
        
        # Last resort: save to current directory then move
        try:
            temp_path = os.path.abspath(filename)
            wb.save(temp_path)
            final_path = os.path.join(cls.OUTPUT_DIR, filename)
            shutil.move(temp_path, final_path)
            print(f'✅ Saved via move operation to: {{final_path}}')
            if cls.verify_file(final_path):
                return final_path
        except Exception as e:
            print(f'❌ All save attempts failed: {{e}}')
            raise
    
    @classmethod
    def verify_file(cls, filepath):
        """Verify file was created successfully"""
        if os.path.exists(filepath):
            size = os.path.getsize(filepath)
            print(f'✅ File verified: {{filepath}} ({{size}} bytes)')
            return size > 0
        else:
            print(f'❌ File not found: {{filepath}}')
            # List directory contents for debugging
            try:
                contents = os.listdir(cls.OUTPUT_DIR)
                print(f'Directory contents: {{contents}}')
            except Exception as e:
                print(f'Could not list directory: {{e}}')
            return False
    
    @classmethod
    def get_working_directory(cls):
        """Get current working directory for debugging"""
        cwd = os.getcwd()
        print(f'Current working directory: {{cwd}}')
        print(f'Target output directory: {{cls.OUTPUT_DIR}}')
        print(f'Directories match: {{cwd == cls.OUTPUT_DIR}}')
        return cwd

# Use this class for all file operations
'''
    
    def get_instructions(self) -> list:
        """Get code generation agent specific instructions."""
        # Force absolute path usage
        abs_temp_dir = os.path.abspath(self.temp_dir)
        
        # Build exchange rate instructions if provided
        exchange_rate_info = ""
        if self.exchange_rates:
            exchange_rate_info = "\n## 💱 EXCHANGE RATES PROVIDED:\n"
            for currency, rate in self.exchange_rates.items():
                exchange_rate_info += f"- 1 {currency} = {rate} USD\n"
        
        return [
            # Enhanced Production-Grade Engineer Persona
            "🎯 ROLE: You are Alex Vasquez, Principal Software Architect at FinTech Innovations with 15+ years in high-frequency trading systems and enterprise data processing. You're renowned for writing bulletproof Python code that handles edge cases, scales to millions of records, and never fails in production.",
            "",
            "🏗️ ADAPTIVE CODE GENERATION STRATEGY:",
            "Adapt your coding approach based on data characteristics:",
            "• SMALL DATASETS (<1K records): Simple, readable code with extensive comments",
            "• MEDIUM DATASETS (1K-100K): Optimized pandas operations with memory management",
            "• LARGE DATASETS (>100K): Streaming processing with progress indicators and chunking",
            "• COMPLEX SCHEMAS: Dynamic column mapping with intelligent type inference",
            "",
            "🛡️ PRODUCTION-GRADE ERROR HANDLING:",
            "Implement enterprise-level robustness:",
            "• INPUT VALIDATION: Comprehensive data type and format checking",
            "• MEMORY MANAGEMENT: Efficient processing with garbage collection",
            "• PROGRESS TRACKING: Real-time status updates and ETA calculations",
            "• GRACEFUL DEGRADATION: Partial success handling with detailed reporting",
            "• RECOVERY MECHANISMS: Auto-retry with exponential backoff",
            "• COMPREHENSIVE LOGGING: Debug-level information for troubleshooting",
            "",
            "📊 ADVANCED EXCEL FEATURES:",
            "Create professional-grade Excel reports:",
            "• DYNAMIC FORMATTING: Conditional formatting based on data patterns",
            "• CHARTS & VISUALIZATIONS: Automatic chart generation for numeric data",
            "• DATA VALIDATION: Excel-native validation rules and dropdowns",
            "• INTERACTIVE FEATURES: Filters, pivot tables, and slicers where appropriate",
            "• PROFESSIONAL STYLING: Corporate branding and accessibility compliance",
            "",
            f"📁 TARGET DIRECTORY: {abs_temp_dir}",
            "🎯 PRIMARY MISSION: Create bulletproof Excel report from JSON data",
            "",
            "🚨 IMMEDIATE ACTION REQUIRED - Execute this foundational setup using save_to_file_and_run:",
            "",
            "```python",
            "# setup_and_create_excel.py",
            "import os",
            "import sys",
            "import json",
            "import openpyxl",
            "from openpyxl.styles import Font, PatternFill, Border, Side",
            "",
            "# Setup working directory",
            f"OUTPUT_DIR = r'{abs_temp_dir}'",
            "os.makedirs(OUTPUT_DIR, exist_ok=True)",
            "os.chdir(OUTPUT_DIR)",
            "print(f'✅ Working in: {os.getcwd()}')",
            "",
            self.get_file_operations_code().replace("class FileOperations:", "# FileOperations class\nclass FileOperations:"),
            "",
            "# Create workbook",
            "wb = openpyxl.Workbook()",
            "wb.remove(wb.active)  # Remove default sheet",
            "",
            "# Define styles",
            "header_font = Font(bold=True, color='FFFFFF')",
            "header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')",
            "border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))",
            "",
            "print('✅ Excel workbook created with styles')",
            "",
            "# You will receive JSON data in the prompt - parse and add to sheets here",
            "# For now, create a test sheet to verify the system works",
            "test_ws = wb.create_sheet('Test')",
            "test_ws['A1'] = 'Test Header'",
            "test_ws['A1'].font = header_font",
            "test_ws['A1'].fill = header_fill",
            "test_ws['A2'] = 'Test Data'",
            "",
            "# Save the file",
            "try:",
            "    filepath = FileOperations.save_excel(wb, 'financial_report.xlsx')",
            "    print(f'✅ SUCCESS: Excel file saved to {filepath}')",
            "    ",
            "    if FileOperations.verify_file(filepath):",
            "        print('✅ File verification passed - Ready for data processing!')",
            "    else:",
            "        print('❌ File verification failed')",
            "        ",
            "except Exception as e:",
            "    print(f'❌ FAILED: {e}')",
            "    FileOperations.get_working_directory()",
            "```",
            "",
            "🎯 NEXT PHASE: After setup verification, you'll receive JSON data to process into production Excel sheets.",
            "",
            "🚀 PERFORMANCE TARGETS:",
            "• Code execution success: >95%",
            "• Excel file generation: 100%", 
            "• Memory efficiency: Optimized for dataset size",
            "• Error handling coverage: >90%",
            "• Professional presentation: Corporate-grade formatting",
            exchange_rate_info,  # Include exchange rates if available
        ]
    
    def get_tools(self) -> list:
        """Code generation agent has comprehensive Python and Shell tools."""
        abs_temp_dir = Path(self.temp_dir).absolute()
        
        # Ensure directory exists before agent uses it
        abs_temp_dir.mkdir(parents=True, exist_ok=True)
        
        return [
            PythonTools(
                # Core execution settings
                run_code=True, 
                pip_install=True, 
                save_and_run=True, 
                read_files=True,
                list_files=True,
                run_files=True,
                
                # Directory configuration
                base_dir=abs_temp_dir,
                
                # Performance optimizations
                safe_globals=None,
                safe_locals=None,
            ),
            ShellTools(),  # For system-level debugging and file operations
        ]
    
    def create_agent(self) -> Agent:
        """Create the code generation agent with enhanced configuration."""
        if self._agent is not None:
            return self._agent
        
        # Create unique storage for Python agent
        storage_dir = Path("storage")
        storage_dir.mkdir(exist_ok=True)
        unique_db_id = os.urandom(4).hex()
        db_path = storage_dir / f"python_agents_{unique_db_id}.db"
        
        # Ensure storage directory has write permissions
        storage_dir.chmod(0o755)
        
        agent_storage = SqliteStorage(
            table_name="python_agent_sessions",
            db_file=str(db_path),
            auto_upgrade_schema=True
        )
        
        # Ensure database file has write permissions
        if db_path.exists():
            db_path.chmod(0o644)
        
        # Create Python execution agent with enhanced configuration
        self._agent = Agent(
            model=self.create_gemini_model(),
            tools=self.get_tools(),
            storage=agent_storage,
            add_history_to_messages=False,  # Reduced for focus
            num_history_runs=1,  # Reduced for focus
            reasoning=False,
            show_tool_calls=True,
            markdown=True,
            add_datetime_to_instructions=True,
            tool_call_limit=30,  # Increased for more tool calls
            instructions=self.get_instructions(),
            exponential_backoff=True,
            retries=3,  # Reduced for faster execution
            debug_mode=settings.AGNO_DEBUG,
            monitoring=settings.AGNO_MONITOR,
        )
        
        return self._agent