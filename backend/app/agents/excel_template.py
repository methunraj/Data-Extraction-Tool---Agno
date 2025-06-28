"""
Template for simple Excel generation from JSON.
This shows the exact pattern we want the agent to follow.
"""

import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def create_clean_excel(json_path: str, output_path: str):
    """
    Create a clean, well-formatted Excel file from JSON data.
    No calculations or analysis - just clean presentation.
    """
    
    # Read JSON data
    with open(json_path, 'r') as f:
        content = f.read()
        
    # Handle markdown-wrapped JSON
    if content.strip().startswith('```json') and content.strip().endswith('```'):
        content = content.strip()[7:-3].strip()
    
    data = json.loads(content)
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # 1. Company Information Sheet
        if 'company_identification' in data:
            company_data = data['company_identification']
            # Flatten nested data
            flat_company = {}
            for key, value in company_data.items():
                if isinstance(value, dict):
                    for sub_key, sub_value in value.items():
                        flat_company[f"{key}_{sub_key}"] = sub_value
                else:
                    flat_company[key] = value
            
            # Create DataFrame with two columns: Field and Value
            company_df = pd.DataFrame([
                {'Field': key.replace('_', ' ').title(), 'Value': str(value)}
                for key, value in flat_company.items()
            ])
            company_df.to_excel(writer, sheet_name='Company Info', index=False)
        
        # 2. Financial Metrics Sheet(s)
        if 'financial_metrics' in data:
            metrics = data['financial_metrics']
            
            if isinstance(metrics, list):
                # All metrics in one sheet
                metrics_df = pd.DataFrame(metrics)
                
                # Flatten any list columns
                for col in metrics_df.columns:
                    if metrics_df[col].apply(lambda x: isinstance(x, list)).any():
                        metrics_df[col] = metrics_df[col].apply(
                            lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x
                        )
                
                # Sort by year if exists
                if 'reporting_year' in metrics_df.columns:
                    metrics_df = metrics_df.sort_values('reporting_year', ascending=False)
                
                metrics_df.to_excel(writer, sheet_name='Financial Metrics', index=False)
            
            elif isinstance(metrics, dict):
                # Separate sheet for each metric category
                for category, items in metrics.items():
                    if isinstance(items, list) and items:
                        df = pd.DataFrame(items)
                        # Flatten lists
                        for col in df.columns:
                            if df[col].apply(lambda x: isinstance(x, list)).any():
                                df[col] = df[col].apply(
                                    lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x
                                )
                        sheet_name = category.replace('_', ' ').title()[:31]
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 3. Extraction Notes Sheet
        if 'extraction_notes' in data:
            notes = data['extraction_notes']
            # Convert to Field/Value format
            notes_data = []
            for key, value in notes.items():
                if isinstance(value, list):
                    value = ', '.join(map(str, value))
                notes_data.append({
                    'Field': key.replace('_', ' ').title(),
                    'Value': str(value)
                })
            notes_df = pd.DataFrame(notes_data)
            notes_df.to_excel(writer, sheet_name='Extraction Notes', index=False)
    
    # Apply formatting
    wb = load_workbook(output_path)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    alternating_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Format each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Format headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Format data rows
        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.border = border
                # Alternating row colors
                if row_num % 2 == 0:
                    cell.fill = alternating_fill
                
                # Number formatting for currency
                if cell.value and isinstance(cell.value, (int, float)):
                    if 'value' in str(ws.cell(1, cell.column).value).lower() or \
                       'amount' in str(ws.cell(1, cell.column).value).lower():
                        cell.number_format = '$#,##0.00'
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze top row
        ws.freeze_panes = 'A2'
    
    # Save formatted workbook
    wb.save(output_path)
    print(f"Excel file created successfully at: {output_path}")


if __name__ == "__main__":
    # Example usage
    create_clean_excel("sample_data.json", "output.xlsx")