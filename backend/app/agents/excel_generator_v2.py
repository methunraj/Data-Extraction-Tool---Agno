"""
Enhanced Excel Generator Agent with retry mechanisms.
Converts JSON data to professional Excel files with better error handling.
"""

import json
import os
import sys
from pathlib import Path
from typing import Optional, Iterator

from agno.agent import Agent
from agno.tools.file import FileTools
from agno.tools.python import PythonTools
from agno.tools import tool
from agno.exceptions import RetryAgentRun


class ExcelGeneratorAgentV2(Agent):
    """
    Enhanced agent specialized in generating Excel files from JSON data.
    Uses custom tools with retry logic to ensure completion.
    """

    def __init__(self, model, working_dir: Optional[str] = None):
        """Initialize Excel generator agent."""

        # Set up working directory
        self.work_dir = working_dir or os.environ.get("AGENT_TEMP_DIR", "/tmp/agno_work")
        Path(self.work_dir).mkdir(parents=True, exist_ok=True)
        
        # Track state
        self.packages_installed = False
        self.json_read = False
        self.excel_created = False

        # Create custom tool for Excel generation
        @tool
        def generate_excel_from_json(
            json_file_path: str, 
            output_excel_path: str,
            install_packages: bool = True
        ) -> Iterator[str]:
            """
            Generate Excel file from JSON data with automatic package installation.
            
            Args:
                json_file_path: Path to input JSON file
                output_excel_path: Path for output Excel file
                install_packages: Whether to install required packages
            """
            yield "Starting Excel generation process..."
            
            # Step 1: Install packages if needed
            if install_packages and not self.packages_installed:
                yield "Installing required packages: pandas and openpyxl..."
                import subprocess
                subprocess.run([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"], check=True)
                self.packages_installed = True
                yield "Packages installed successfully!"
            
            # Step 2: Import packages
            try:
                import pandas as pd
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils.dataframe import dataframe_to_rows
                yield "Imported required libraries"
            except ImportError as e:
                raise RetryAgentRun(f"Failed to import packages: {e}. Please install pandas and openpyxl first.")
            
            # Step 3: Read JSON data
            yield f"Reading JSON data from: {json_file_path}"
            try:
                with open(json_file_path, 'r') as f:
                    data = json.load(f)
                self.json_read = True
                yield "JSON data loaded successfully"
            except Exception as e:
                raise RetryAgentRun(f"Failed to read JSON file: {e}")
            
            # Step 4: Create Excel with multiple sheets
            yield "Creating Excel workbook..."
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Sheet 1: Company Information
            if 'company_identification' in data:
                ws1 = wb.create_sheet("Company Information")
                company_data = data['company_identification']
                
                # Flatten company data
                rows = []
                for key, value in company_data.items():
                    rows.append([key.replace('_', ' ').title(), str(value)])
                
                for row in rows:
                    ws1.append(row)
                
                # Format headers
                for cell in ws1['A']:
                    cell.font = Font(bold=True)
                
                yield "Created Company Information sheet"
            
            # Sheet 2: Financial Metrics
            if 'financial_metrics' in data:
                metrics = data['financial_metrics']
                
                # Create sheets for each metric type
                for metric_type, metric_data in metrics.items():
                    if isinstance(metric_data, list) and metric_data:
                        sheet_name = metric_type.replace('_', ' ').title()[:31]  # Excel sheet name limit
                        ws = wb.create_sheet(sheet_name)
                        
                        # Get all unique keys from the metric data
                        all_keys = set()
                        for item in metric_data:
                            if isinstance(item, dict):
                                all_keys.update(item.keys())
                        
                        # Create header
                        headers = list(all_keys)
                        ws.append(headers)
                        
                        # Add data
                        for item in metric_data:
                            row = [item.get(key, '') for key in headers]
                            ws.append(row)
                        
                        # Format headers
                        for cell in ws[1]:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                        
                        # Auto-adjust column widths
                        for column in ws.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            ws.column_dimensions[column_letter].width = adjusted_width
                        
                        yield f"Created {sheet_name} sheet"
            
            # Sheet 3: Metadata
            if 'extraction_metadata' in data:
                ws_meta = wb.create_sheet("Metadata")
                metadata = data['extraction_metadata']
                
                rows = []
                for key, value in metadata.items():
                    rows.append([key.replace('_', ' ').title(), str(value)])
                
                for row in rows:
                    ws_meta.append(row)
                
                # Format headers
                for cell in ws_meta['A']:
                    cell.font = Font(bold=True)
                
                yield "Created Metadata sheet"
            
            # Step 5: Save Excel file
            yield f"Saving Excel file to: {output_excel_path}"
            try:
                wb.save(output_excel_path)
                self.excel_created = True
                yield f"Excel file created successfully at: {output_excel_path}"
            except Exception as e:
                raise RetryAgentRun(f"Failed to save Excel file: {e}")
            
            # Verify file exists
            if os.path.exists(output_excel_path):
                file_size = os.path.getsize(output_excel_path)
                yield f"✓ Excel file verified: {file_size} bytes"
                return f"Successfully created Excel file at {output_excel_path}"
            else:
                raise RetryAgentRun("Excel file was not created. Retrying...")

        # Initialize with custom tool and standard tools
        super().__init__(
            name="ExcelGeneratorV2",
            model=model,
            tools=[
                generate_excel_from_json,
                PythonTools(
                    base_dir=Path(self.work_dir),
                    pip_install=True,
                    run_code=True,
                    save_and_run=True,
                    run_files=True,
                    read_files=True,
                ),
                FileTools(base_dir=Path(self.work_dir)),
            ],
            description="You are an automated Excel generation specialist with a custom tool that handles the entire JSON-to-Excel conversion process efficiently and reliably.",
            goal="Use the generate_excel_from_json tool to create a professional Excel file from the provided JSON data path.",
            instructions=[
                "Call generate_excel_from_json with the exact input JSON file path and output Excel file path provided in the prompt",
                "Monitor the tool's progress output carefully",
                "If the tool reports any errors, analyze them and retry if needed",
                "If you encounter 'string indices must be integers' or similar errors:",
                "  - Use read_file to examine the JSON file structure",
                "  - Use run_python_code to test JSON parsing",
                "  - Debug the issue step by step",
                "  - Write a custom Python script if the tool fails",
                "Ensure the Excel file is created successfully at the specified location",
                "You have access to Python tools for debugging and custom solutions",
            ],
            expected_output="Successful execution of generate_excel_from_json tool resulting in a complete Excel file at the specified location with all JSON data properly organized into sheets.",
            additional_context="The generate_excel_from_json tool automatically handles package installation, JSON parsing, sheet creation, and formatting. It will create sheets for company information, financial metrics, and metadata.",
            stream=False,
            show_tool_calls=True,
            add_history_to_messages=True,
            num_history_responses=10,
            exponential_backoff=True,
            retries=20,
            markdown=False,
        )